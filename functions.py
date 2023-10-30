# Selenium library
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, WebDriverException

# Other library
import pandas as pd
from textwrap import wrap
import calendar
from functools import reduce
from datetime import datetime
from datetime import date
from dateutil.relativedelta import relativedelta
import copy
import time

# Excel libraries
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.chart import BarChart3D, Reference, BarChart


# functions for scraping

def second_fridays(year):
    # get second friday of the each for each month
    c = calendar.Calendar(firstweekday=calendar.SUNDAY)
    second_fri = []
    for m in range(1,13):
        fri = c.monthdatescalendar(year, m)[1][5]
        second_fri.append(fri)
    return second_fri

def get_next_month():
    # get current year and month 
    current_year = date.today().year  
    current_month_of_option = date.today().month

    # get second friday of the month when option expires
    second_fridays_of_month = second_fridays(current_year)[current_month_of_option-1]

    # if (date.today()>second_fridays_of_month):
    #     second_fridays_of_month = second_fridays_of_month + relativedelta(months=+1)
    # else:
    #     pass
    
    # if today is after the second friday of the current month, the option of the next month is expired,
    # get the after the next month
    if (date.today()>second_fridays_of_month):
        next_month = date.today() + relativedelta(months=+2)
    else:
        next_month = date.today() + relativedelta(months=+1)
        
    next_month = next_month.strftime('%h').lower()
    return next_month

def fix_header(driver):
    
    shadow_host = driver.find_elements_by_xpath("//bc-data-grid[@data-ng-hide='error']")[0]
    shadow_root = shadow_host.shadow_root
    shadow_content = shadow_root.find_elements(By.CLASS_NAME, 'bc-datatable-header-tooltip')
    headers_list = []
    for i in range(len(shadow_content)):
        headers_list.append(shadow_content[i].text)
    headers = headers_list[:10]
    return headers
    
def fix_header_2(driver):
    header = driver.find_element_by_xpath('//thead')
    header = header.text.replace('\n', ' ').split()
    headers = []
    headers.append(header[0])
    headers_end = ' '.join(header[1:3])
    headers.append(headers_end)
    headers_end = ' '.join(header[3:5])
    headers.append(headers_end)    
    headers_end = ' '.join(header[5:7])
    headers.append(headers_end)    
    headers.append(header[7])
    headers_end = ' '.join(header[8:10])
    headers.append(headers_end)
    headers_end = ' '.join(header[10:12])
    headers.append(headers_end)
    headers_end = ' '.join(header[12:14])
    headers.append(headers_end)
    headers.extend(header[14:-1])
    return headers

def check_month(month_abb):
    list_month = calendar.month_abbr[1:]
    list_month = [month.lower() for month in list_month]
    dict_symbol = {'jul': 'KCN',
               'sep': 'KCU',
               'dec': 'KCZ',
               'mar': 'KCH',
               'may': 'KCK'}
    if month_abb in dict_symbol:
        return dict_symbol[month_abb]
    else:
        idx = list_month.index(month_abb)
        idx_counter = 1
        while True:
            if list_month[idx+idx_counter] in dict_symbol:
                return dict_symbol[list_month[idx+idx_counter]]
            else:
                
                idx_counter+=1

def scrape_header_and_body(current_symbol, current_year, next_month):

    # start scrape the initial page
    options = Options()
    options.page_load_strategy = 'normal'
    driver = webdriver.Chrome('/Users/domenicospoto/Downloads/chromedriver', options=options)
    driver.get('https://www.barchart.com/futures/quotes/'+current_symbol+current_year+'/options/'+next_month+'-'+current_year+'?futuresOptionsTime=daily&moneyness=20')
    time.sleep(7)

    result = None
    while result is None:
        try:
            driver.find_element_by_xpath("//button[@class='Button__StyledButton-a1qza5-0 cUAUIG']").click()
            result = 'Done'
            print('1st obstacle overcame.')
        except WebDriverException:    
            print('1st obstacle.')

    result = None
    while result is None:
        try:
            driver.find_element_by_xpath("//button[@class='Button__StyledButton-a1qza5-0 eofJul']").click()
            result = 'Done'
            print('2st obstacle overcame.')
        except WebDriverException:    
            print('2st obstacle.')



    print('All OK. Start scraping ...')

    # find name of financial instrument
    instrument = driver.find_element_by_class_name('symbol') 
    instrument = instrument.text.split()[0]

    # find today date
    today_date = driver.find_element_by_xpath("//span[@class='current-date ng-binding']").text[-14:]
    today_date = datetime.strptime(today_date, 
                                 "%b %dst, %Y")

    # find expiration date
    # expiration_date = driver.find_element_by_xpath("//div[@class='column small-12 medium-4']").text

    # find month and year available of options
    dates = driver.find_element_by_id('bc-options-toolbar__dropdown-month')  
    dates = wrap(dates.text.replace('\n', ' ').strip().lower(),8)  # replace space with dash, then cut space, then lower string, then get the month and year
    dates = [date.replace(' ', '-') for date in dates]
    dates = [d[:4] + d[6:] for d in dates]

    table = dict()
    for index, date in enumerate(dates):

        if index == 0:

            # # get the daily prices
            # driver.find_element_by_xpath("//option[@value='daily']").click()
            # # get all the strike prices
            # driver.find_element_by_xpath("//option[@value='allRows']").click()

            headers = fix_header(driver)

            # find body of table
            shadow_host = driver.find_elements_by_xpath("//bc-data-grid[@data-ng-hide='error']")[0]
            shadow_root = shadow_host.shadow_root
            shadow_content = shadow_root.find_elements(By.CSS_SELECTOR, "text-binding")

            rows_grid = [i for i in range(0, len(shadow_content)+1, 10)]
                
            body = list()

            for idx in range(1, len(rows_grid)):
                
                line = []
                for i in range(rows_grid[idx-1], rows_grid[idx]):
                    line.append(shadow_content[i].text)
                body.append(line)

            print('Body scraped for : {}'.format(date))
            
            table[date] = pd.DataFrame(body, columns=headers)

            # get currect strike price
            current_strike = driver.find_element_by_xpath("//span[@class='last-change ng-binding']").text[:-1]
        
        elif index == 4:
            print('Body scraped ended')
            break
            
        else: 

            symbol = check_month(date[:3])
            year = date[-2:]
            driver.get('https://www.barchart.com/futures/quotes/'+symbol+year+'/options/'+date+'?futuresOptionsTime=daily&moneyness=20')

            time.sleep(8)

            # find header of table
            headers = None
            while headers is None:
                try:
                    headers = fix_header(driver)
                    print('Headers loaded from page.')
                except:
                    print('Not possible to load headers from page.')
            

            # find body of table
            shadow_host = driver.find_elements_by_xpath("//bc-data-grid[@data-ng-hide='error']")[0]
            shadow_root = shadow_host.shadow_root
            shadow_content = shadow_root.find_elements(By.CSS_SELECTOR, "text-binding")

            rows_grid = [i for i in range(0, len(shadow_content)+1, 10)]
                
            body = list()

            for idx in range(1, len(rows_grid)):
                
                line = []
                for i in range(rows_grid[idx-1], rows_grid[idx]):
                    line.append(shadow_content[i].text)
                body.append(line)

            print('Body scraped for : {}'.format(date))

            table[date] = pd.DataFrame(body, columns=headers)

            # # find body of table
            # counter = 0
            # body = list()
            # while True:

            #     try:
            #         line = driver.find_element_by_xpath('//tr[@data-current-index='+str(counter)+']')
            #         line = line.text.replace('\n', ' ').split()
            #         body.append(line)
            #         counter +=1
            #     except:
            #         print('Body scraped for : {}'.format(date))
            #         break
            
            # table[date] = pd.DataFrame(body, columns=headers)

            # get currect strike price
            #current_strikes[date] = driver.find_element_by_xpath("//span[@class='last-change ng-binding']").text

    driver.close()

    print('WebPage closed. Scraping ended.')

    return instrument, today_date, current_strike, table

def build_table(table, today_date, current_strike):
    # adjust dates for the front months considered
    dates = [date for date in table.keys()]

    # Build table of Strike and Open Int
    open_int_dict = dict()
    for date in dates: 
        open_int_dict[date] = table[date].loc[:, ['Strike','Open Interest']]   # select only interested col
        open_int_dict[date]['Open Interest'] = open_int_dict[date]['Open Interest'].str.replace(',', '')  # replace , to . in open int
        open_int_dict[date]['Open Interest'] = pd.to_numeric(open_int_dict[date]['Open Interest'], errors='coerce') # convert to number Open Int str
        open_int_dict[date]['Open Interest'] = open_int_dict[date]['Open Interest'].fillna(0)
        open_int_dict[date].set_index('Strike', inplace=True)             # place Strike as Index

    open_int_list = [open_int_dict[date] for date in dates]   # put in List each dataframe
    open_int = reduce(lambda x,y: pd.merge(x,y, on='Strike'), open_int_list) # combine all of them in one single df
    open_int.columns = dates   # set col as dates

    # Build table for Strike and Open Int
    open_int = open_int.reset_index()                            # reset index with number

    # for Call Options
    open_int_call_ = open_int[open_int.Strike.str.contains('C')]  # select only strikes of Call Opt
    open_int_call_['Strike'] = pd.to_numeric(open_int_call_['Strike'].apply(lambda x: x.replace('C', ''))) # cut out the C
    open_int_call_ = open_int_call_.reset_index(drop=True)      # reset again to numbers the index
    open_int_call_1 = open_int_call_[(open_int_call_['Strike'] - float(current_strike))<0][-10:]  # get the small closest 10 strike prices w.r.t. current strike
    open_int_call_2 = open_int_call_[(open_int_call_['Strike'] - float(current_strike))>0][:10]  # get the big closest 10 strike prices w.r.t. current strike
    open_int_call = pd.concat([open_int_call_1, open_int_call_2])
    open_int_call.set_index('Strike', inplace=True)
    open_int_call = open_int_call.T
    open_int_call.reset_index(inplace=True)
    open_int_call = open_int_call.rename({'index': 'Date'}, axis=1)
    # today_date = datetime.strptime(today_date, '%m/%d/%y')
    open_int_call['Date'] = today_date

    # for Put Options
    open_int_put_ = open_int[open_int.Strike.str.contains('P')]  # select only strikes of Call Opt
    open_int_put_['Strike'] = pd.to_numeric(open_int_put_['Strike'].apply(lambda x: x.replace('P', ''))) # cut out the C
    open_int_put_ = open_int_put_.reset_index(drop=True)      # reset again to numbers the index
    open_int_put_1 = open_int_put_[(open_int_put_['Strike'] - float(current_strike))<0][-10:]  # get the small closest 10 strike prices w.r.t. current strike
    open_int_put_2 = open_int_put_[(open_int_put_['Strike'] - float(current_strike))>0][:10]  # get the big closest 10 strike prices w.r.t. current strike
    open_int_put = pd.concat([open_int_put_1, open_int_put_2])
    open_int_put.set_index('Strike', inplace=True)
    open_int_put = open_int_put.T
    open_int_put.reset_index(inplace=True)
    open_int_put = open_int_put.rename({'index': 'Date'}, axis=1)
    open_int_put['Date'] = today_date

    return open_int_call, open_int_put, dates

# functions for creating Excel and tables
def table_call_options(open_int_call, ws):

    # For Call Option Part
    ft = Font(name='Arial', size=12, bold=True)
    ws['K4'].font = ft
    ws['K4'] = 'CALL'
    red = Color(rgb='49b600')
    fill = PatternFill(fill_type='solid', fgColor=red)
    ws['K4'].fill = fill

    open_int_call_copy = copy.deepcopy(pd.DataFrame(open_int_call))
    open_int_call_copy.reset_index(inplace=True)
    open_int_call_copy.columns = ['Strike', 'Open Int']
    
    for index, entry in open_int_call_copy.iterrows():
        ws.cell(row=5, column=1+index, value=entry['Strike'])
        ws.cell(row=6, column=1+index, value=entry['Open Int'])
    
    ft = Font(name='Arial',
        size=12, 
        bold=True, 
        underline='single')
    for col in ws.iter_cols(min_col=1, max_col=len(open_int_call_copy)+1, min_row=5, max_row=5):
        for cell in col:
            cell.font = ft

    ft = Font(name='Arial',
        size=12)
    for col in ws.iter_cols(min_col=1, max_col=len(open_int_call_copy)+1, min_row=6, max_row=6):
        for cell in col:
            cell.font = ft

    data = Reference(ws, min_col=2, min_row=5, max_col=len(open_int_call_copy), max_row=6)
    titles = Reference(ws, min_col=1, min_row=6) # , max_row=6
    #chart = BarChart3D()
    chart = BarChart()
    chart.title = "Call Open  Int vs. Strike per date"
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)
    chart.height = 17
    chart.width = 20

    ws.add_chart(chart, "D21")


def table_put_options(open_int_put, ws):

    # For Put Option Part
    ft = Font(name='Arial', size=12, bold=True)
    ws['AH4'].font = ft    # WARNING: it is a static table, it should change w.r.t. the number of front months
    ws['AH4'] = 'PUT'     # WARNING: it is a static table, it should change w.r.t. the number of front months
    red = Color(rgb='ff726f')
    fill = PatternFill(fill_type='solid', fgColor=red)
    ws['AH4'].fill = fill   # WARNING: it is a static table, it should change w.r.t. the number of front months

    open_int_put_copy = copy.deepcopy(pd.DataFrame(open_int_put))
    open_int_put_copy.reset_index(inplace=True)
    open_int_put_copy.columns = ['Strike', 'Open Int']
    
    for index, entry in open_int_put_copy.iterrows():
        ws.cell(row=5, column=24+index, value=entry['Strike'])
        ws.cell(row=6, column=24+index, value=entry['Open Int'])
    
    ft = Font(name='Arial',
        size=12, 
        bold=True, 
        underline='single')
    for col in ws.iter_cols(min_col=len(open_int_put_copy)+2, max_col=len(open_int_put_copy)+23, min_row=5, max_row=5):
        for cell in col:
            cell.font = ft

    ft = Font(name='Arial',
        size=12)
    for col in ws.iter_cols(min_col=len(open_int_put_copy)+2, max_col=len(open_int_put_copy)+23, min_row=6, max_row=6):
        for cell in col:
            cell.font = ft

    data = Reference(ws, min_col=25, min_row=5, max_col=len(open_int_put_copy)+23, max_row=6)
    titles = Reference(ws, min_col=24, min_row=6) # , max_row=6
    #chart = BarChart3D()
    chart = BarChart()
    chart.title = "Put Open  Int vs. Strike per date"
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)
    chart.height = 17
    chart.width = 20

    ws.add_chart(chart, "AA21")

def update_table_call_options(open_int_call, ws, day):

    open_int_call_copy = copy.deepcopy(list(open_int_call))

    for index, entry in enumerate(open_int_call_copy, start=1):
        ws.cell(row=5+day, column=index, value=entry)

    ft = Font(name='Arial',
        size=12)

    for col in ws.iter_cols(min_col=1, max_col=len(open_int_call_copy)+1, min_row=5+day, max_row=6+day):
        for cell in col:
            cell.font = ft

    # delete previous plot
    del ws._charts[0]
    
    data = Reference(ws, min_col=2, min_row=5, max_col=len(open_int_call_copy), max_row=7)
    titles = Reference(ws, min_col=1, min_row=6, max_row=6+day-1)
    # chart = BarChart3D()
    chart = BarChart()
    chart.title = "Call Open  Int vs. Strike per date"
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)
    chart.height = 15 
    chart.width = 20 

    ws.add_chart(chart, "D21")

def update_table_put_options(open_int_put, ws, day):

    open_int_put_copy = copy.deepcopy(list(open_int_put))

    for index, entry in enumerate(open_int_put_copy, start=1):
        ws.cell(row=5+day, column=23+index, value=entry)

    ft = Font(name='Arial',
        size=12)

    for col in ws.iter_cols(min_col=len(open_int_put_copy)+2, max_col=len(open_int_put_copy)+23, min_row=5+day, max_row=6+day):
        for cell in col:
            cell.font = ft
    
    # delete previous plot
    del ws._charts[0]

    data = Reference(ws, min_col=25, min_row=5, max_col=len(open_int_put_copy)+23, max_row=7)
    titles = Reference(ws, min_col=24, min_row=6, max_row=6+day-1)
    # chart = BarChart3D()
    chart = BarChart()
    chart.title = "Put Open  Int vs. Strike per date"
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)
    chart.height = 15 
    chart.width = 20 

    ws.add_chart(chart, "AA21")


def table_call_options_all(open_int_call, ws, dates):

    # For Call Option Part
    ft = Font(name='Arial', size=12, bold=True)
    ws['K4'].font = ft
    ws['K4'] = 'CALL'
    red = Color(rgb='49b600')
    fill = PatternFill(fill_type='solid', fgColor=red)
    ws['K4'].fill = fill

    for i in range(len(open_int_call)):

        if i == 0:

            open_int_call_copy = copy.deepcopy(pd.DataFrame(open_int_call.loc[i]))
            open_int_call_copy.reset_index(inplace=True)
            open_int_call_copy.columns = ['Strike', 'Open Int']
    
            for index, entry in open_int_call_copy.iterrows():
                if index == 0:
                    ws.cell(row=5, column=1+index, value=entry['Strike'])
                    ws.cell(row=6, column=1+index, value=dates[i])                    
                else:
                    ws.cell(row=5, column=1+index, value=entry['Strike'])
                    ws.cell(row=6, column=1+index, value=entry['Open Int'])
    
            ft = Font(name='Arial',
                size=12, 
                bold=True, 
                underline='single')
            for col in ws.iter_cols(min_col=1, max_col=len(open_int_call_copy)+1, min_row=5, max_row=5):
                for cell in col:
                    cell.font = ft

            ft = Font(name='Arial',
                size=12)
            for col in ws.iter_cols(min_col=1, max_col=len(open_int_call_copy)+1, min_row=6, max_row=6):
                for cell in col:
                    cell.font = ft

        else:
            open_int_call_copy = copy.deepcopy(list(open_int_call.loc[i]))

            for index, entry in enumerate(open_int_call_copy, start=1):
                if index == 1:
                    ws.cell(row=5+i+1, column=index, value=dates[i])
                else:
                    ws.cell(row=5+i+1, column=index, value=entry)

            ft = Font(name='Arial',
                size=12)

            for col in ws.iter_cols(min_col=1, max_col=len(open_int_call_copy)+1, min_row=5+i+1, max_row=6+i+1):
                for cell in col:
                    cell.font = ft



    data = Reference(ws, min_col=2, min_row=5, max_col=len(open_int_call_copy), max_row=9)
    titles = Reference(ws, min_col=1, min_row=6, max_row=9) # , max_row=6
    #chart = BarChart3D()
    chart = BarChart()
    chart.title = "Call Open  Int vs. Strike per date"
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)
    chart.height = 17
    chart.width = 20

    ws.add_chart(chart, "D21")

def table_put_options_all(open_int_put, ws, dates):

    # For Put Option Part
    ft = Font(name='Arial', size=12, bold=True)
    ws['AH4'].font = ft    # WARNING: it is a static table, it should change w.r.t. the number of front months
    ws['AH4'] = 'PUT'     # WARNING: it is a static table, it should change w.r.t. the number of front months
    red = Color(rgb='ff726f')
    fill = PatternFill(fill_type='solid', fgColor=red)
    ws['AH4'].fill = fill   # WARNING: it is a static table, it should change w.r.t. the number of front months

    for i in range(len(open_int_put)):

        if i == 0:

            open_int_put_copy = copy.deepcopy(pd.DataFrame(open_int_put.loc[i]))
            open_int_put_copy.reset_index(inplace=True)
            open_int_put_copy.columns = ['Strike', 'Open Int']
    
            for index, entry in open_int_put_copy.iterrows():
                if index == 0:
                    ws.cell(row=5, column=24+index, value=entry['Strike'])
                    ws.cell(row=6, column=24+index, value=dates[i])
                else:
                    ws.cell(row=5, column=24+index, value=entry['Strike'])
                    ws.cell(row=6, column=24+index, value=entry['Open Int'])                    
            
            ft = Font(name='Arial',
                size=12, 
                bold=True, 
                underline='single')
            for col in ws.iter_cols(min_col=len(open_int_put_copy)+2, max_col=len(open_int_put_copy)+23, min_row=5, max_row=5):
                for cell in col:
                    cell.font = ft

            ft = Font(name='Arial',
                size=12)
            for col in ws.iter_cols(min_col=len(open_int_put_copy)+2, max_col=len(open_int_put_copy)+23, min_row=6, max_row=6):
                for cell in col:
                    cell.font = ft

        else:

            open_int_put_copy = copy.deepcopy(list(open_int_put.loc[i]))

            for index, entry in enumerate(open_int_put_copy, start=1):
                if index == 1:
                    ws.cell(row=5+i+1, column=23+index, value=dates[i])
                else:
                    ws.cell(row=5+i+1, column=23+index, value=entry)

            ft = Font(name='Arial',
                size=12)

            for col in ws.iter_cols(min_col=len(open_int_put_copy)+2, max_col=len(open_int_put_copy)+23, min_row=5+i+1, max_row=6+i+1):
                for cell in col:
                    cell.font = ft

    data = Reference(ws, min_col=25, min_row=5, max_col=len(open_int_put_copy)+23, max_row=9)
    titles = Reference(ws, min_col=24, min_row=6, max_row=9) # , max_row=6
    #chart = BarChart3D()
    chart = BarChart()
    chart.title = "Put Open  Int vs. Strike per date"
    chart.add_data(data=data, titles_from_data=True)
    chart.set_categories(titles)
    chart.height = 17
    chart.width = 20

    ws.add_chart(chart, "AA21")