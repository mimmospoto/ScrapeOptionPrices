# Other library
from datetime import date
from datetime import date
import os

# Excel libraries
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# Our library
from functions import *

# *************************************************************************************
# CHANGE THE LINE BELOW WITH YOUR PATH FOLDER WHERE YOU WANT TO SAVE THE EXCEL FILE

path = '/Users/domenicospoto/Desktop/Project_Upwork/DerekPanaia/'+instrument+'.xlsx'

# *************************************************************************************


# *************************************************************************************
# CHANGE THE LINE BELOW WITH THE NUMBER OF DAY YOU RUN THE CODE
# i.e. if it is the  first time you run the code, put day = 1. If it is the second day
# you run the code, put day = 2

day = 1

# *************************************************************************************


# get next month for scraping algorithm at the first page
next_month = get_next_month()

# get current year (only last 2 digits) for scraping algorithm at the first page
current_year = str(date.today().year)[2:]

# get current symbol name for scraping algorithm at the first page
current_symbol = check_month(next_month)

print('Date: {}'.format(date.today()))

# scrape the hrader and body for the strike dates needed
instrument, today_date, current_strike, table = scrape_header_and_body(current_symbol, current_year, next_month)

# build table for call and put options
open_int_call, open_int_put, dates = build_table(table, today_date, current_strike)

today_date_ = str(today_date.year)+'_'+str(today_date.day)+'_'+str(today_date.month)

# Create Excel and tables
if os.path.exists(path):

    # IF EXCEL FILE EXISTS, IT WILL RUN THIS
    # UPDATE THE EXCEL FILE STARTING FROM DAY 2 OF SCRAPING DATA
    wb = load_workbook('Coffee.xlsx')

    for idx, date in enumerate(dates):
        ws = wb[date]

        update_table_call_options(open_int_call.loc[idx], ws, day=day) # CHANGE THE DAY
        update_table_put_options(open_int_put.loc[idx], ws, day=day)   # CHANGE THE DAY

    # FOR PLOT OF ALL EXPIRATION DATE TOGETHER
    ws = wb.create_sheet('ALL_'+today_date_)
    ws.title = 'ALL_' + today_date_
    ft = Font(name='Arial', size=14, bold=True)
    ws['A1'].font = ft
    ws['A1'] = ws.title

    table_call_options_all(open_int_call, ws, dates)
    table_put_options_all(open_int_put, ws, dates)

else:

    # IF EXCEL FILE NOT EXIST, IT WILL RUN THIS
    # IT CREATES THE FIRST DAY OF VALUES AND THE EXCEL FILE
    wb = Workbook()

    for idx, date in enumerate(dates):
        ws = wb.create_sheet(date)
        ws.title = date

        ft = Font(name='Arial', size=14, bold=True)
        ws['A1'].font = ft
        ws['A1'] = ws.title

        table_call_options(open_int_call.loc[idx], ws)
        table_put_options(open_int_put.loc[idx], ws)
    
    # wb.remove_sheet('Sheet')
    # del wb['Sheet']

    # FOR PLOT OF ALL EXPIRATION DATE TOGETHER
    ws = wb.create_sheet('ALL_'+today_date_)
    ws.title = 'ALL_' + today_date_
    ft = Font(name='Arial', size=14, bold=True)
    ws['A1'].font = ft
    ws['A1'] = ws.title

    table_call_options_all(open_int_call, ws, dates)
    table_put_options_all(open_int_put, ws, dates)



wb.save(instrument+'.xlsx')

